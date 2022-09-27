from flask import Flask, render_template, url_for, redirect, flash, send_file
from mainApp import app, db
from mainApp.forms import AddQuestionType1Form, AddQuestionType2Form,AssessmentForm
from mainApp.models import QuestionTypeOne, QuestionType2,Asessment,questions
from flask_wtf import FlaskForm
from flask_sqlalchemy import SQLAlchemy
from wtforms_sqlalchemy.fields import QuerySelectField
from openpyxl import load_workbook



@app.route("/")
@app.route("/home")
def home():
    return render_template('home.html',title='Home')

@app.route("/staff")
def staff():
    type1questions=QuestionTypeOne.query.all()
    return render_template('staff.html', title='Staff', type1questions=type1questions)

@app.route("/student")
def student():
    return render_template('student.html',title='Student')    

@app.route("/add_question_type1", methods=['POST', 'GET'])
def add_question_type1():
    form = AddQuestionType1Form()
    if form.validate_on_submit():
        form_answer_options = form.answer_options.data.split(',')
        form_answer_option_1 = form_answer_options[0]
        form_answer_option_2 = form_answer_options[1]
        form_answer_option_3 = form_answer_options[2]

        form_question_tags = form.question_tags.data.split(',')
        form_question_tag_1 = form_question_tags[0]
        form_question_tag_2 = form_question_tags[1]
        form_question_tag_3 = form_question_tags[2]

        question = QuestionTypeOne(question=form.question.data, answer_option_1=form_answer_option_1, answer_option_2=form_answer_option_2, answer_option_3=form_answer_option_3, correct_answer=form.correct_answer.data, marks_available=form.marks_available.data, question_tag_1=form_question_tag_1, question_tag_2=form_question_tag_2, question_tag_3=form_question_tag_3, correct_answer_feedback=form.correct_answer_feedback.data, incorrect_answer_feedback=form.incorrect_answer_feedback.data, feedforward_comments=form.feedforward_comments.data)
        db.session.add(question)
        db.session.commit()
        flash('Question added.')
        return redirect(url_for('staff'))
    return render_template('add_question_type1.html',title='Add Multiple Choice Question', form=form)

@app.route("/add_text_question", methods=['GET','POST'])
def addtextquestion():
    form = AddQuestionType2Form()
    if form.validate_on_submit():
        question2 = QuestionType2(questionType = 2, name = form.name.data, shortDescription = form.shortDescription.data, question = form.question.data, answer = form.answer.data, correctFeedback = form.correctFeedback.data, incorrectFeedback = form.incorrectFeedback.data, marksAwarded = form.marksAwarded.data)
        db.session.add(question2)
        db.session.commit()
        return redirect(url_for('staff'))
    return render_template('addq2.html', title = 'Add Text Question', form=form)

@app.route("/addaquestion")
def addaquestion():
    return render_template('questiontype.html')

@app.route('/temp_question_2_list')
def question2List():
    type2Questions = QuestionType2.query.all()
    return render_template('temp_question_list.html',type2Questions=type2Questions) 

@app.route('/question_type_2/<int:question_id>')
def questionType2(question_id):
    type2Questions = QuestionType2.query.get_or_404(question_id)
    return render_template('question_type_2.html', type2Questions=type2Questions)

@app.route('/question_type_2/edit/<int:question_id>', methods=['GET','POST'])
def edit_question_2(question_id):
    type2Questions = QuestionType2.query.get_or_404(question_id)
    form = AddQuestionType2Form()
    if form.validate_on_submit():
        type2Questions.name = form.name.data
        type2Questions.shortDescription = form.shortDescription.data
        type2Questions.question = form.question.data
        type2Questions.answer = form.answer.data
        type2Questions.correctFeedback = form.correctFeedback.data
        type2Questions.incorrectFeedback = form.incorrectFeedback.data
        type2Questions.marksAwarded = form.marksAwarded.data
        db.session.add(type2Questions)
        db.session.commit()
        flash('Successfully updated post')
        return redirect(url_for('question2List'))
    form.name.data = type2Questions.name
    form.shortDescription.data = type2Questions.shortDescription
    form.question.data = type2Questions.question
    form.answer.data = type2Questions.answer
    form.correctFeedback.data = type2Questions.correctFeedback
    form.incorrectFeedback.data = type2Questions.incorrectFeedback
    form.marksAwarded.data = type2Questions.marksAwarded
    return render_template('edit_questions_2.html', form=form, type2Questions=type2Questions)


@app.route('/question_type_2/delete/<int:question_id>')
def delete_question_type_2(question_id):
    question_to_delete = QuestionType2.query.get_or_404(question_id)
    db.session.delete(question_to_delete)
    db.session.commit()
    flash('Successfully deleted question.')
    return redirect(url_for('question2List'))

@app.route("/edit_type_one_question/<id>", methods=['POST', 'GET'])
def edit_type_one_question(id):
    type_one_question = QuestionTypeOne.query.filter_by(id=id).first()
    form = AddQuestionType1Form()
    if form.validate_on_submit():
        form_answer_options = form.answer_options.data.split(",")
        form_answer_option_1 = form_answer_options[0]
        form_answer_option_2 = form_answer_options[1]
        form_answer_option_3 = form_answer_options[2]

        form_question_tags = form.question_tags.data.split(",")
        form_question_tag_1 = form_question_tags[0]
        form_question_tag_2 = form_question_tags[1]
        form_question_tag_3 = form_question_tags[2]
        
        type_one_question.question = form.question.data
        type_one_question.answer_option_1 = form_answer_option_1
        type_one_question.answer_option_2 = form_answer_option_2
        type_one_question.answer_option_3 = form_answer_option_3
        type_one_question.correct_answer = form.correct_answer.data
        type_one_question.marks_available = form.marks_available.data
        type_one_question.question_tag_1 = form_question_tag_1
        type_one_question.question_tag_2 = form_question_tag_2
        type_one_question.question_tag_3 = form_question_tag_3
        type_one_question.correct_answer_feedback = form.correct_answer_feedback.data
        type_one_question.incorrect_answer_feedback = form.incorrect_answer_feedback.data
        type_one_question.feedforward_comments = form.feedforward_comments.data
        db.session.add(type_one_question)
        db.session.commit()
        flash('Question updated.')
        return redirect(url_for('home'))
    form.question.data = type_one_question.question 
    form.answer_options.data = f"{type_one_question.answer_option_1}, {type_one_question.answer_option_2}, {type_one_question.answer_option_3}"
    form.correct_answer.data = type_one_question.correct_answer
    form.marks_available.data = type_one_question.marks_available
    form.question_tags.data = f"{type_one_question.question_tag_1}, {type_one_question.question_tag_2}, {type_one_question.question_tag_3}"
    form.correct_answer_feedback.data = type_one_question.correct_answer_feedback
    form.incorrect_answer_feedback.data = type_one_question.incorrect_answer_feedback
    form.feedforward_comments.data = type_one_question.feedforward_comments

    
    return render_template('edit_type_one_question.html',title='Edit Multiple Choice Question', form=form, type_one_question=type_one_question)

@app.route("/delete_type_one_question/<id>")
def delete_type_one_question(id):
    type_one_question = QuestionTypeOne.query.filter_by(id=id).first()
    if not type_one_question:
        flash("Question does not exist", category="error")
    else:
        db.session.delete(type_one_question)
        db.session.commit()
        flash("Question deleted", category="success")
    return redirect(url_for('home'))


# assessment

@app.route("/assessment_create", methods=["GET","POST"])
def assessment_create():
    form = AssessmentForm()
    if form.validate_on_submit():
        title = form.assessment_title.data
        kind =  int(form.assessment_type.data)
        difficulty = int(form.assessment_difficulty.data)
        model = Asessment(assessment_title=title,assessment_difficulty=difficulty,assessment_type=kind)
        db.session.add(model)
        db.session.commit()
        flash("assessment created Successfully. Please add the question by clicking one assessment",category="success")
        return redirect(url_for('assessment_home'))
    return render_template('assessment_create.html',form=form)




@app.route("/assessment_home",methods=["GET","POST"])
def assessment_home():
    query = Asessment.query.all()
    return render_template("assessment_home.html",query=query)


@app.route('/assessment_view/<int:id>',methods=["GET","POST"])
def assessment_view(id):
    model = Asessment.query.filter_by(id=id).first()
    return render_template("assessment_view.html",model=model)





@app.route('/assessment/question/add/<int:id>/',methods=["GET","POST"])
def assessment_question_add(id):
    ass = Asessment.query.filter_by(id=id).first()

    if ass.assessment_type == 0:
        questions = QuestionTypeOne.query.all()
    else:
        questions = QuestionType2.query.all()


    return render_template("form.html",questions=questions,ass=id)


@app.route('/upload_questiom/<int:id>/<int:ass>')
def upload_question(id,ass):
    k = QuestionTypeOne.query.filter_by(id=id).first()
    i = questions(question=k.question,assessment_id=ass)
    db.session.add(i)
    db.session.commit()
    flash("question add successfull")

    return redirect(url_for('assessment_question_add',id=ass))





@app.route("/assessment_edit/<int:id>",methods=["GET","POST"])
def assessment_edit(id):
    model = Asessment.query.filter_by(id=id).first()
    form = AssessmentForm(obj=model)
    if form.validate_on_submit():
        form.populate_obj(model)
        db.session.add(model)
        db.session.commit()
        return redirect(url_for('assessment_main'))
    return render_template("assessment.html",form=form) 


@app.route('/assessment_delete/<int:id>',methods=["GET","POST"])
def assessment_delete(id):
    model = Asessment.query.filter_by(id=id).first()
    db.session.delete(model)
    db.session.commit()
    flash("Successfully Deleted")
    return redirect(url_for('assessment_home'))


@app.route("/review_statistics")
def review_statistics ():
    # book1 = load_workbook("data1.xlsx")
    # sheet = book1.active
    book2 = load_workbook("data2.xlsx")
    sheet = book2.active
    return render_template("review_statistics.html", sheet=sheet)

@app.route("/download1")
def download_image ():
    path = "table1.jpg"
    return send_file(path, as_attachment=True)

@app.route("/download2")
def download_table ():
    path = "data2.xlsx"
    return send_file(path, as_attachment=True)